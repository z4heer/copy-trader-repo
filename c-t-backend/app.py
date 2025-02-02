from flask import Flask, request, jsonify
from flask_cors import CORS
import os
import logging
import traceback
from openpyxl import load_workbook
from flask import send_from_directory, make_response

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Configure logging
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s [%(levelname)s] %(message)s',
                    handlers=[
                        logging.FileHandler("app.log"),
                        logging.StreamHandler()
                    ])

# Dummy data for demonstration
active_users = [
    {"apiKey": "xxx", "api_secret_password": "xxx", "reqId": "1", "userid": "U1001", "active": True, "quantity": 10},
    {"apiKey": "xxx", "api_secret_password": "xxx", "reqId": "2", "userid": "U1002", "active": True, "quantity": 20},
    {"apiKey": "xxx", "api_secret_password": "xxx", "reqId": "3", "userid": "U1003", "active": False, "quantity": 15}
]

orders = [
  {
    "orderId": "B1",
    "type": "Buy",
    "quantity": 10,
    "status": "pending",
    "timestamp": "2025-01-30T10:15:00Z"
  },
  {
    "orderId": "B2",
    "type": "Buy",
    "quantity": 20,
    "status": "completed",
    "timestamp": "2025-01-29T14:30:00Z"
  },
  {
    "orderId": "M1",
    "type": "Modify",
    "quantity": 15,
    "status": "pending",
    "timestamp": "2025-01-30T11:00:00Z"
  },
  {
    "orderId": "C1",
    "type": "Cancel",
    "quantity": 5,
    "status": "cancelled",
    "timestamp": "2025-01-28T09:45:00Z"
  }
]
  # In-memory list to hold orders
net_positions = [
  {
    "positionId": "NP1",
    "symbol": "AAPL",
    "quantity": 50,
    "pnl": 120.5,
    "lastUpdated": "2025-01-30T10:00:00Z"
  },
  {
    "positionId": "NP2",
    "symbol": "GOOGL",
    "quantity": 30,
    "pnl": -50.0,
    "lastUpdated": "2025-01-29T15:00:00Z"
  },
  {
    "positionId": "NP3",
    "symbol": "MSFT",
    "quantity": 40,
    "pnl": 75.2,
    "lastUpdated": "2025-01-30T09:30:00Z"
  }
]
  # In-memory net positions (for demonstration)

# Error handler
@app.errorhandler(Exception)
def handle_exception(e):
    logging.error("Unhandled Exception: %s", traceback.format_exc())
    return jsonify({"error": "An internal error occurred."}), 500


def read_active_users():
    """
    Reads the users.xlsx file from the conf folder and returns only the active users.
    Assumes the Excel file has a header row with the following columns:
    apiKey, api_secret_password, reqId, userid, active, quantity
    """
    file_path = os.path.join('conf', 'users.xlsx')

    # Load the workbook and select the active sheet
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active

    # Read header row (first row)
    header = [cell.value for cell in sheet[1]]
    active_users = []

    # Iterate through the remaining rows
    for row in sheet.iter_rows(min_row=2, values_only=True):
        user = dict(zip(header, row))
        # Filter active users. Adjust the condition if 'active' is stored as a boolean or a string.
        if user.get('active') in [True, 'TRUE', 'True']:
            active_users.append(user)

    return active_users


@app.route('/active-users', methods=['GET'])
def get_active_users():
    """
    Flask endpoint to return active users read from the Excel configuration file.
    """
    try:
        users = read_active_users()
        logging.info("Fetched %d active users.", len(users))
        return jsonify(users), 200
    except Exception as e:
        logging.error("Error reading users file: %s", str(e))
        return jsonify({"error": "Unable to read user config file"}), 500


# Endpoint: Place Buy Order
@app.route('/buy-order', methods=['POST'])
def place_buy_order():
    try:
        # Here you would iterate through active users and call the Nuvama Wealth API
        # For now, we simply simulate order placement.
        order = {"orderId": "B" + str(len(orders)+1), "status": "pending", "details": "Buy order placed for active users"}
        orders.append(order)
        logging.info("Buy order placed: %s", order)
        return jsonify(order), 200
    except Exception as e:
        logging.error("Error in place_buy_order: %s", str(e))
        return jsonify({"error": "Failed to place buy order"}), 500

# Endpoint: List Orders
@app.route('/orders', methods=['GET'])
def list_orders():
    try:
        logging.info("Listing all orders")
        return jsonify(orders), 200
    except Exception as e:
        logging.error("Error in list_orders: %s", str(e))
        return jsonify({"error": "Failed to list orders"}), 500

# Endpoint: Modify Order
@app.route('/modify-order/<orderId>', methods=['PUT'])
def modify_order(orderId):
    try:
        update_data = request.get_json()
        # Find the order and update it (dummy logic)
        for order in orders:
            if order["orderId"] == orderId:
                order.update(update_data)
                logging.info("Order %s modified: %s", orderId, update_data)
                return jsonify(order), 200
        logging.warning("Order %s not found", orderId)
        return jsonify({"error": "Order not found"}), 404
    except Exception as e:
        logging.error("Error in modify_order: %s", str(e))
        return jsonify({"error": "Failed to modify order"}), 500

# Endpoint: Cancel Order
@app.route('/cancel-order/<orderId>', methods=['DELETE'])
def cancel_order(orderId):
    try:
        global orders
        orders = [order for order in orders if order["orderId"] != orderId]
        logging.info("Order %s cancelled", orderId)
        return jsonify({"message": "Order cancelled"}), 200
    except Exception as e:
        logging.error("Error in cancel_order: %s", str(e))
        return jsonify({"error": "Failed to cancel order"}), 500

# Endpoint: List Net Positions
@app.route('/net-positions', methods=['GET'])
def list_net_positions():
    try:
        logging.info("Listing net positions")
        return jsonify(net_positions), 200
    except Exception as e:
        logging.error("Error in list_net_positions: %s", str(e))
        return jsonify({"error": "Failed to fetch net positions"}), 500

# Endpoint: Sell Order
@app.route('/sell-order', methods=['POST'])
def sell_order():
    try:
        data = request.get_json()
        orderId = data.get("orderId")
        # Dummy logic to simulate selling an order from net positions
        sell_result = {"orderId": orderId, "status": "sold", "details": "Sell order executed"}
        logging.info("Sell order executed: %s", sell_result)
        return jsonify(sell_result), 200
    except Exception as e:
        logging.error("Error in sell_order: %s", str(e))
        return jsonify({"error": "Failed to execute sell order"}), 500

@app.route('/favicon.ico')
def favicon():
    # Option A: Return a 204 No Content response
    return make_response('', 204)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)